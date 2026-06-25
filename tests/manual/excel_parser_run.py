#!/usr/bin/env python3
"""
ExcelParser 解析测试
使用本地文件 /Users/zhangzheng/zavixai/4G5G吞吐量数据.xlsx 验证：
- 表头包含“统计月份”
- 数据行的“统计月份”列不为空，且为日期文本

运行：
python3 /Users/zhangzheng/zavixai/Sage/tests/test_excel_parser.py
"""

import os
import sys
import re

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")))


def main():
    print("📈 ExcelParser 解析测试")
    print("=" * 60)

    excel_path = "/Users/zhangzheng/zavixai/4G5G吞吐量数据.xlsx"
    if not os.path.exists(excel_path):
        print(f"❌ 文件不存在: {excel_path}")
        return False

    # 尝试导入解析器（避开包级依赖，直接从文件路径加载）
    try:
        import importlib.util

        module_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "sagents",
            "tool",
            "file_parser_tool.py",
        )
        spec = importlib.util.spec_from_file_location("file_parser_tool", module_path)
        file_parser_tool = importlib.util.module_from_spec(spec)  # pyright: ignore[reportArgumentType]
        spec.loader.exec_module(file_parser_tool)  # pyright: ignore[reportOptionalMemberAccess]
        ExcelParser = file_parser_tool.ExcelParser
    except Exception as e:
        print(f"❌ 加载 ExcelParser 模块失败: {e}")
        print("💡 请确保已安装依赖: pip install openpyxl docstring_parser")
        return False

    try:
        md_text, meta = ExcelParser.extract_text_from_xlsx(excel_path)
        print("✅ 解析成功，生成Markdown长度:", len(md_text))
        sheets = meta.get("sheets", [])
        print("📋 工作表:", sheets)
        if not sheets:
            print("❌ 未检测到工作表")
            return False

        # 检查表头是否包含“统计月份”
        has_header = False
        for sn in sheets:
            headers = meta.get("sheet_" + sn, {}).get("headers", [])
            print(f"🧾 {sn} 表头: {headers}")
            if "统计月份" in headers:
                has_header = True
        if not has_header:
            print("❌ 所有工作表的表头都不包含‘统计月份’")
            return False

        # 检查Markdown中是否包含日期文本（yyyy-mm-dd）
        date_pattern = re.compile(r"\b\d{4}-\d{2}-\d{2}\b")
        date_matches = date_pattern.findall(md_text)
        print("🗓️ 检测到日期样本数量:", len(date_matches))
        if len(date_matches) == 0:
            print("❌ Markdown中未检测到日期文本，可能‘统计月份’列为空或未文本化")
            return False

        # 进一步使用 openpyxl 验证单元格值（如果可用）
        try:
            from openpyxl import load_workbook

            wb = load_workbook(excel_path, data_only=True, read_only=False)
            for sname in wb.sheetnames:
                sd = wb[sname]
                max_row, max_col = sd.max_row, sd.max_column
                # 寻找表头所在行（第一行）
                header = [str(sd.cell(1, c).value or "") for c in range(1, max_col + 1)]
                if "统计月份" in header:
                    col_idx = header.index("统计月份") + 1
                    sample_vals = []
                    for r in range(2, min(max_row + 1, 12)):
                        v = sd.cell(r, col_idx).value
                        sample_vals.append(v)
                    print(f"🧪 {sname} ‘统计月份’样本值:", sample_vals)
                    if all(v is None or str(v).strip() == "" for v in sample_vals):
                        print("❌ openpyxl 读取到的‘统计月份’列均为空")
                        wb.close()
                        return False
            wb.close()
        except Exception as e:
            # openpyxl 不可用不影响主要测试，通过Markdown与元数据已验证
            print(f"⚠️ openpyxl 验证步骤跳过: {e}")

        print("\n🎉 测试通过：‘统计月份’列解析正常且非空")
        return True

    except Exception as e:
        print(f"❌ 解析过程中发生异常: {e}")
        import traceback

        traceback.print_exc()
        return False


if __name__ == "__main__":
    try:
        ok = main()
        print("\n" + ("🎊 演示完成！" if ok else "💥 演示失败！"))
        sys.exit(0 if ok else 1)
    except KeyboardInterrupt:
        print("\n⏹️ 用户中断测试")
        sys.exit(0)
    except Exception as e:
        print(f"\n💥 测试程序异常: {e}")
        sys.exit(1)
