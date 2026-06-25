#!/usr/bin/env python3
import os
import sys

EXCEL_PATH = "/Users/zhangzheng/zavixai/4G5G吞吐量数据.xlsx"


def main():
    print("🔎 检查 Excel ‘统计月份’列的原始值与公式值")
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ 文件不存在: {EXCEL_PATH}")
        return 1

    try:
        from openpyxl import load_workbook
    except Exception as e:
        print(f"❌ openpyxl 导入失败: {e}")
        return 1

    try:
        wb_data = load_workbook(EXCEL_PATH, data_only=True, read_only=False)
        wb_formula = load_workbook(EXCEL_PATH, data_only=False, read_only=False)
    except Exception as e:
        print(f"❌ 打开工作簿失败: {e}")
        return 1

    for sname in wb_data.sheetnames:
        ws_data = wb_data[sname]
        ws_formula = wb_formula[sname]
        max_row, max_col = ws_data.max_row, ws_data.max_column
        header = [
            str(ws_data.cell(1, c).value or "").strip() for c in range(1, max_col + 1)
        ]
        print(f"🧾 工作表: {sname} 表头: {header}")
        if "统计月份" not in header:
            continue
        col_idx = header.index("统计月份") + 1
        data_vals = []
        formula_vals = []
        for r in range(2, min(max_row + 1, 22)):
            dv = ws_data.cell(r, col_idx).value
            fv = ws_formula.cell(r, col_idx).value
            data_vals.append(dv)
            formula_vals.append(fv)
        print(f"📊 data_only 值样本: {data_vals}")
        print(f"🧮 公式文本样本: {formula_vals}")
    wb_data.close()
    wb_formula.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
