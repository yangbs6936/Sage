#!/usr/bin/env python3
"""
使用 openpyxl 验证 /Users/zhangzheng/zavixai/4G5G吞吐量数据.xlsx：
- 表头包含“统计月份”
- 数据行“统计月份”列非空（存在日期或字符串）

运行示例：
  conda run -n zz python3 /Users/zhangzheng/zavixai/Sage/tests/test_excel_openpyxl_only.py
"""

import os
import sys

EXCEL_PATH = "/Users/zhangzheng/zavixai/4G5G吞吐量数据.xlsx"


def main() -> bool:
    print("📊 openpyxl 解析测试：统计月份列验证")
    print("=" * 60)

    if not os.path.exists(EXCEL_PATH):
        print(f"❌ 文件不存在: {EXCEL_PATH}")
        return False

    try:
        from openpyxl import load_workbook
    except Exception as e:
        print(f"❌ openpyxl 未安装或导入失败: {e}")
        print("💡 请在zz环境中安装: pip install openpyxl")
        return False

    try:
        wb = load_workbook(EXCEL_PATH, data_only=True, read_only=False)
    except Exception as e:
        print(f"❌ 打开工作簿失败: {e}")
        return False

    try:
        success_any = False
        for sname in wb.sheetnames:
            ws = wb[sname]
            max_row, max_col = ws.max_row, ws.max_column
            # 假设第一行为表头
            header = [
                str(ws.cell(1, c).value or "").strip() for c in range(1, max_col + 1)
            ]
            print(f"🧾 工作表: {sname} 表头: {header}")

            if "统计月份" not in header:
                continue

            col_idx = header.index("统计月份") + 1
            # 采样前10行（第2行开始）
            sample = []
            for r in range(2, min(max_row + 1, 12)):
                v = ws.cell(r, col_idx).value
                sample.append(v)
            print(f"🧪 ‘统计月份’样本值: {sample}")

            if any(v is not None and str(v).strip() != "" for v in sample):
                success_any = True
                break

        wb.close()

        if success_any:
            print("\n🎉 验证通过：‘统计月份’列存在非空值")
            return True
        else:
            print("\n❌ 验证失败：未在任何工作表中检测到‘统计月份’非空值")
            return False

    except Exception as e:
        print(f"❌ 解析过程异常: {e}")
        return False


if __name__ == "__main__":
    ok = main()
    print("\n" + ("🎊 演示完成！" if ok else "💥 演示失败！"))
    sys.exit(0 if ok else 1)
